import re, datetime as dt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# ====== Texto do calendário (FIFA - PT) ======
# Fonte: calendário oficial divulgado pela FIFA. [1](https://www.fifa.com/pt/tournaments/mens/worldcup/canadamexicousa2026/articles/copa-mundo-2026-tabela-jogos)
src = r'''
Quinta-feira, 11 de junho de 2026Quinta-feira, 11 de junho de 2026
Grupo A: México x África do Sul Cidade do México, no México – 13h00 no horário local (16h00 em Brasília / 18h00 em Praia / 20h00 em Lisboa)
Grupo A: República da Coreia x República Tcheca Guadalajara – 20h00 no horário local (23h00 em Brasília / 1h00 de 12 de junho em Praia / 3h00 de 12 de junho em Lisboa)
Sexta-feira, 12 de junho de 2026Sexta-feira, 12 de junho de 2026
Grupo B: Canadá x Bósnia e Herzegovina Toronto, no Canadá – 15h00 no horário local (16h00 em Brasília / 18h00 em Praia / 20h00 em Lisboa)
Grupo D: Estados Unidos x Paraguai Los Angeles, nos EUA – 18h00 no horário local (22h00 em Brasília / 0h00 de 13 de junho em Praia / 2h00 de 13 de junho em Lisboa)
Sábado, 13 de junho de 2026Sábado, 13 de junho de 2026
Grupo B: Catar x Suíça Santa Clara (região da Baía de San Francisco), nos EUA – 12h00 no horário local (16h00 em Brasília / 20h00 em Praia / 22h00 em Lisboa)
Grupo C: Brasil x Marrocos Nova York/Nova Jersey, nos EUA – 18h00 no horário local (19h00 em Brasília / 21h00 em Praia / 23h00 em Lisboa)
Grupo C: Haiti x Escócia Boston, nos EUA – 21h00 no horário local (22h00 em Brasília / 0h00 de 14 de junho em Praia / 2h00 de 14 de junho em Lisboa)
Grupo D: Austrália x Turquia Vancouver, no Canadá – 21h00 no horário local (1h00 de 14 de junho em Brasília / 3h00 em Praia / 5h00 em Lisboa)
Domingo, 14 de junho de 2026Domingo, 14 de junho de 2026
Grupo E: Alemanha x Curaçau Houston, nos EUA – 12h00 no horário local (14h00 em Brasília / 16h00 em Praia / 18h00 em Lisboa)
Grupo E: Costa do Marfim x Equador Filadélfia, nos EUA – 19h00 no horário local (20h00 em Brasília / 22h00 em Praia / 0h00 de 15 de junho em Lisboa)
Grupo F: Holanda x Japão Dallas, nos EUA – 15h00 no horário local (17h00 em Brasília / 19h00 em Praia / 21h00 em Lisboa)
Grupo F: Suécia x Tunísia Monterrey, no México – 20h00 no horário local (23h00 em Brasília / 1h00 de 15 de junho em Praia / 3h00 de 15 de junho em Lisboa)
Segunda-feira, 15 de junho de 2026Segunda-feira, 15 de junho de 2026
Grupo H: Espanha x Cabo Verde Atlanta, nos EUA – 12h00 no horário local (13h00 em Brasília / 15h00 em Praia / 17h00 em Lisboa)
Grupo H: Arábia Saudita x Uruguai Miami, nos EUA – 18h00 no horário local (19h00 em Brasília / 21h00 em Praia / 23h00 em Lisboa)
Grupo G: Bélgica x Egito Seattle, nos EUA – 12h00 no horário local (16h00 em Brasília / 18h00 em Praia / 20h00 em Lisboa)
Grupo G: Irã x Nova Zelândia Los Angeles, nos EUA – 18h00 no horário local (22h00 em Brasília / 0h00 de 16 de junho em Praia / 2h00 de 16 de junho em Lisboa)
Terça-feira, 16 de junho de 2026Terça-feira, 16 de junho de 2026
Grupo J: Áustria x Jordânia Santa Clara (região da Baía de San Francisco), nos EUA – 21h00 no horário local (1h00 de 17 de junho em Brasília / 3h00 de 17 de junho em Praia / 5h00 de 17 de junho em Lisboa)
Grupo I: França x Senegal Nova York/Nova Jersey, nos EUA – 15h00 no horário local (16h00 em Brasília / 18h00 em Praia / 20h00 em Lisboa)
Grupo I: Iraque x Noruega Boston, nos EUA – 18h00 no horário local (19h00 em Brasília / 21h00 em Praia / 23h00 em Lisboa)
Grupo J: Argentina x Argélia Kansas City, nos EUA – 20h00 no horário local (22h00 em Brasília / 0h00 de 17 de junho em Praia / 2h00 de 17 de junho em Lisboa)
Quarta-feira, 17 de junho de 2026Quarta-feira, 17 de junho de 2026
Grupo K: Portugal x República Democrática do Congo Houston, nos EUA – 12h00 no horário local (14h00 em Brasília / 16h00 em Praia / 18h00 em Lisboa)
Grupo L: Inglaterra x Croácia Dallas, nos EUA – 15h00 no horário local (17h00 em Brasília / 19h00 em Praia / 21h00 em Lisboa)
Grupo L: Gana x Panamá Toronto, no Canadá – 19h00 no horário local (20h00 em Brasília / 22h00 em Praia / 0h00 de 18 de junho em Lisboa)
Grupo K: Uzbequistão x Colômbia Cidade do México, no México – 20h00 no horário local (21h00 em Brasília / 23h00 em Praia / 1h00 de 18 de junho em Lisboa)

Quinta-feira, 18 de junho de 2026Quinta-feira, 18 de junho de 2026
Grupo A: República Tcheca x África do Sul Atlanta, nos EUA – 12h00 no horário local (13h00 em Brasília / 15h00 em Praia / 17h00 em Lisboa)
Grupo B: Suíça x Bósnia e Herzegovina Los Angeles, nos EUA – 12h00 no horário local (16h00 em Brasília / 18h00 em Praia / 20h00 em Lisboa)
Grupo B: Canadá x Catar Vancouver, no Canadá – 15h00 no horário local   (19h00 em Brasília / 21h00 em Praia / 23h00 em Lisboa)
Grupo A: México x República da Coreia Guadalajara, no México – 19h00 no horário local (22h00 em Brasília / 0h00 de 19 de junho em Praia / 2h00 de 19 de junho em Lisboa)
Sexta-feira, 19 de junho de 2026Sexta-feira, 19 de junho de 2026
Grupo D: Turquia x Paraguai Santa Clara (região da Baía de San Francisco), nos EUA – 20h00 no horário local (0h00 em Brasília / 2h00 em Praia / 4h00 em Lisboa)
Grupo D: Estados Unidos x Austrália Seattle, nos EUA – 12h00 no horário local (16h00 em Brasília / 18h00 em Praia / 20h00 em Lisboa)
Grupo C: Escócia x Marrocos Boston, nos EUA – 18h00 no horário local (19h00 em Brasília / 21h00 em Praia / 23h00 em Lisboa)
Grupo C: Brasil x Haiti Filadélfia, nos EUA – 20h30 no horário local (21h30 em Brasília / 23h30 de 20 de junho em Praia / 1h30 de 20 de junho em Lisboa)
Sábado 20 de junho de 2026Sábado 20 de junho de 2026
Grupo F: Tunísia x Japão Monterrey, no México – 20h00 no horário local (23h00 em Brasília / 1h00 de 21 de junho em Praia/ 3h00 de 21 de junho em Lisboa)
Grupo F: Holanda x Suécia Houston, nos EUA – 12h00 no horário local (14h00 em Brasília / 16h00 em Praia / 18h00 em Lisboa)
Grupo E: Alemanha x Costa do Marfim Toronto, no Canadá – 16h00 no horário local (17h00 em Brasília / 19h00 em Praia / 21h00 em Lisboa)
Grupo E: Equador x Curaçau Kansas City, nos EUA – 19h00 no horário local (21h00 em Brasília / 23h00 em Praia / 1h00 de 21 de junho em Lisboa)

Domingo, 21 de janeiro de 2026Domingo, 21 de janeiro de 2026
Grupo H: Espanha x Arábia Saudita Atlanta, nos EUA – 12h00 no horário local (13h00 em Brasília / 15h00 em Praia / 17h00 em Lisboa)
Grupo G: Bélgica x Irã Los Angeles, nos EUA – 12h00 no horário local (16h00 em Brasília / 18h00 em Praia / 20h00 em Lisboa)
Grupo H: Uruguai x Cabo Verde Miami, nos EUA – 18h00 no horário local (19h00 em Brasília / 21h00 em Praia / 23h00 em Lisboa)
Grupo G: Nova Zelândia x Egito Vancouver, no Canadá – 18h00 no horário local (22h00 em Brasília / 0h00 de 22 de junho em Praia / 2h00 de 22 de junho em Lisboa)
Segunda-feira, 22 de janeiro de 2026Segunda-feira, 22 de janeiro de 2026
Grupo J: Argentina x Áustria Dallas, nos EUA – 12h00 no horário local (14h00 em Brasília / 16h00 em Praia / 18h00 em Lisboa)
Grupo I: França x Iraque Filadélfia, nos EUA – 17h00 no horário local (18h00 em Brasília / 20h00 em Praia / 22h00 em Lisboa)
Grupo I: Noruega x Senegal Nova York/Nova Jersey, nos EUA – 20h00 no horário local (21h00 em Brasília / 23h00 em Praia / 1h00 de 23 de junho em Lisboa)
Grupo J: Jordânia x Argélia Santa Clara (região da Baía de San Francisco), nos EUA – 20h00 no horário local (0h00 de 23 de junho em Brasília / 2h00 de 23 de junho em Praia / 4h00 de 23 de junho em Lisboa)
Terça-feira, 23 de junho de 2026Terça-feira, 23 de junho de 2026
Grupo K: Portugal x Uzbequistão Houston, nos EUA – 12h00 no horário local (14h00 em Brasília / 16h00 em Praia / 18h00 em Lisboa)
Grupo L: Inglaterra x Gana Boston, nos EUA – 16h00 no horário local (17h00 em Brasília / 19h00 em Praia / 21h00 em Lisboa)
Grupo L: Panamá x Croácia Toronto, no Canadá – 19h00 no horário local (20h00 em Brasília / 22h00 em Praia / 0h00 de 24 de junho em Lisboa)
Grupo K: Colômbia x República Democrática do Congo Guadalajara, no México – 20h00 no horário local (23h00 em Brasília / 1h00 de 24 de junho em Praia / 3h00 de 24 de junho em Lisboa)

Quarta-feira, 24 de junho de 2026Quarta-feira, 24 de junho de 2026
Grupo B: Suíça x Canadá Vancouver, no Canadá – 12h00 no horário local (16h00 em Brasília / 18h00 em Praia / 20h00 em Lisboa)
Grupo B: Bósnia e Herzegovina x Catar Seattle, nos EUA – 12h00 no horário local (16h00 em Brasília / 18h00 em Praia / 20h00 em Lisboa)
Grupo C: Escócia x Brasil Miami, nos EUA – 18h00 no horário local (19h00 em Brasília / 21h00 em Praia / 23h00 em Lisboa)
Grupo C: Marrocos x Haiti Atlanta, nos EUA – 18h00 no horário local (19h00 em Brasília / 21h00 em Praia / 23h00 em Lisboa)
Grupo A: República Tcheca x México Cidade do México, no México – 19h00 no horário local (22h00 em Brasília / 0h00 de 25 de junho em Praia / 2h00 de 25 de junho em Lisboa)
Grupo A: África do Sul x República da Coreia Monterrey, no México – 19h00 no horário local (22h00 em Brasília / 0h00 de 25 de junho em Praia / 2h00 de 25 de junho em Lisboa)
Quinta-feira, 25 de junho de 2026Quinta-feira, 25 de junho de 2026
Grupo E: Equador x Alemanha Nova York/Nova Jersey, nos EUA – 16h00 no horário local (17h00 em Brasília / 19h00 em Praia / 21h00 em Lisboa)
Grupo E: Curaçau x Costa do Marfim Filadélfia, nos EUA – 16h00 no horário local (17h00 em Brasília / 19h00 em Praia / 21h00 em Lisboa)
Grupo F: Japão x Suécia Dallas, nos EUA – 18h00 no horário local (20h00 em Brasília / 22h00 em Praia / 0h00 de 26 de junho em Lisboa)
Grupo F: Tunísia x Holanda Kansas City, nos EUA – 18h00 no horário local (20h00 em Brasília / 22h00 em Praia / 0h00 de 26 de junho em Lisboa)
Grupo D: Turquia x Estados Unidos Los Angeles, nos EUA – 19h00 no horário local (23h00 em Brasília / 1h00 de 26 de junho em Praia / 3h00 de 26 de junho em Lisboa)
Grupo D: Paraguai x Austrália Santa Clara (região da Baía de San Francisco), nos EUA – 19h00 no horário local (23h00 em Brasília / 1h00 de 26 de junho em Praia / 3h00 de 26 de junho em Lisboa)
Sexta-feira, 26 de junho de 2026Sexta-feira, 26 de junho de 2026
Grupo I: Noruega x França Boston, nos EUA – 15h00 no horário local (16h00 em Brasília / 18h00 em Praia / 20h00 em Lisboa)
Grupo I: Senegal x Iraque Toronto, no Canadá – 15h00 no horário local (16h00 em Brasília / 18h00 em Praia / 20h00 em Lisboa)
Grupo H: Cabo Verde x Arábia Saudita Houston, nos EUA – 19h00 no horário local (21h00 em Brasília / 23h00 em Praia / 1h00 de 27 de junho em Lisboa)
Grupo H: Uruguai x Espanha Guadalajara, no México – 18h00 no horário local (21h00 em Brasília / 23h00 em Praia / 1h00 de 27 de junho em Lisboa)
Grupo G: Egito x Irã Seattle, nos EUA – 20h00 no horário local (0h00 de 27 de junho em Brasília / 2h00 de 27 de junho em Praia / 4h00 de 27 de junho em Lisboa)
Grupo G: Nova Zelândia x Bélgica Vancouver, no Canadá – 20h00 no horário local (0h00 de 27 de junho em Brasília / 2h00 de 27 de junho em Praia / 4h00 de 27 de junho em Lisboa)
Sábado, 27 de junho de 2026Sábado, 27 de junho de 2026
Grupo L: Panamá x Inglaterra Nova York/Nova Jersey, nos EUA – 17h00 no horário local (18h00 em Brasília / 20h00 em Praia / 22h00 em Lisboa)
Grupo L: Croácia x Gana Filadélfia, nos EUA – 17h00 no horário local (18h00 em Brasília / 20h00 em Praia / 22h00 em Lisboa)
Grupo K: Colômbia x Portugal Miami, nos EUA – 19h30 no horário local (20h30 em Brasília / 22h30 em Praia / 0h30 de 28 de junho em Lisboa)
Grupo K: República Democrática do Congo x Uzbequistão Atlanta, nos EUA – 19h30 no horário local (20h30 em Brasília / 22h30 em Praia / 0h30 de 28 de junho em Lisboa)
Grupo J: Argélia x Áustria Kansas City, nos EUA – 21h00 no horário local (23h00 em Brasília / 1h00 de 28 de junho em Praia / 3h00 de 28 de junho em Lisboa)
Grupo J: Jordânia x Argentina Dallas, nos EUA – 21h00 no horário local (23h00 de 28 de junho em Brasília / 0h00 de 28 de junho em Praia / 2h00 de 28 de junho em Lisboa)

Domingo, 28 de junho de 2026Domingo, 28 de junho de 2026
Jogo 73 - Segundo colocado do Grupo A x Segundo colocado do Grupo B - Los Angeles, nos EUA
Segunda-feira, 29 de junho de 2026Segunda-feira, 29 de junho de 2026
Jogo 74 - Vencedor do Grupo E x Terceiro colocado dos Grupos A/B/C/D/F - Boston, nos EUA
Jogo 75 - Vencedor do Grupo F x Segundo colocado do Grupo C - Monterrey, no México
Jogo 76 - Vencedor do Grupo C x Segundo colocado do Grupo F - Houston, nos EUA
Terça-feira, 30 de junho de 2026Terça-feira, 30 de junho de 2026
Jogo 77 - Vencedor do Grupo I x Terceiro colocado dos Grupos C/D/F/G/H - Nova York/Nova Jersey, nos EUA
Jogo 78 - Segundo colocado do Grupo E x Segundo colocado do Grupo I - Dallas, nos EUA
Jogo 79 - Vencedor do Grupo A x Terceiro colocado dos Grupos C/E/F/H/I - Cidade do México, no México
Quarta-feira, 1º de julho de 2026Quarta-feira, 1º de julho de 2026
Jogo 80 - Vencedor do Grupo L x Terceiro colocado dos Grupos E/H/I/J/K - Atlanta, nos EUA
Jogo 81 - Vencedor do Grupo D x Terceiro colocado dos Grupos B/E/F/I/J - Santa Clara (região da Baía de San Francisco), nos EUA
Jogo 82 - Vencedor do Grupo G x Terceiro colocado dos Grupos A/E/H/I/J - Seattle, nos EUA
Quinta-feira, 2 de julho de 2026Quinta-feira, 2 de julho de 2026
Jogo 83 - Segundo colocado do Grupo K x Segundo colocado do Grupo L - Toronto, no Canadá
Jogo 84 - Vencedor do Grupo H x Segundo colocado do Grupo J - Los Angeles, nos EUA
Jogo 85 - Vencedor do Grupo B x Terceiro colocado dos Grupos E/F/G/I/J - Vancouver, no Canadá
Sexta-feira, 3 de julho de 2026Sexta-feira, 3 de julho de 2026
Jogo 86 - Vencedor do Grupo J x Segundo colocado do Grupo H - Miami, nos EUA
Jogo 87 - Vencedor do Grupo K x Terceiro colocado dos Grupos D/E/I/J/L - Kansas City, nos EUA
Jogo 88 - Segundo colocado do Grupo D x Segundo colocado do Grupo G - Dallas, nos EUA
Sábado, 4 de julho de 2026Sábado, 4 de julho de 2026
Jogo 89 – Vencedor do jogo 74 x Vencedor do jogo 77 - Filadélfia, nos EUA
Jogo 90 – Vencedor do jogo 73 x Vencedor do jogo 75 - Houston, nos EUA
Domingo, 5 de julho de 2026Domingo, 5 de julho de 2026
Jogo 91 - Vencedor do jogo 76 x Vencedor do jogo 78 - Nova York/Nova Jersey, nos EUA
Jogo 92 - Vencedor do jogo 79 x Vencedor do jogo 80 - Cidade do México, no México
Segunda-feira, 6 de julho de 2026Segunda-feira, 6 de julho de 2026
Jogo 93 - Vencedor do jogo 83 x Vencedor do jogo 84 - Dallas, nos EUA
Jogo 94 - Vencedor do jogo 81 x Vencedor do jogo 82 - Seattle, nos EUA
Terça-feira, 7 de julho de 2026Terça-feira, 7 de julho de 2026
Jogo 95 – Vencedor do jogo 86 x Vencedor do jogo 88 - Atlanta, nos EUA
Jogo 96 – Vencedor do jogo 85 x Vencedor do jogo 87 - Vancouver, no Canadá
Quinta-feira, 9 de julho de 2026Quinta-feira, 9 de julho de 2026
Jogo 97 - Vencedor do jogo 89 x Vencedor do jogo 90 - Boston, nos EUA
Sexta-feira, 10 de julho de 2026Sexta-feira, 10 de julho de 2026
Jogo 98 - Vencedor do jogo 93 x Vencedor do jogo 94 - Los Angeles, nos EUA
Sábado, 12 de julho de 2026Sábado, 12 de julho de 2026
Jogo 99 - Vencedor do jogo 91 x Vencedor do jogo 92 - Miami, nos EUA
Jogo 100 - Vencedor do jogo 95 x Vencedor do jogo 96 - Kansas City, nos EUA
Terça-feira, 14 de julho de 2026Terça-feira, 14 de julho de 2026
Jogo 101 - Vencedor do jogo 97 x Vencedor do jogo 98 - Dallas, nos EUA
Quarta-feira, 15 de julho 2026Quarta-feira, 15 de julho 2026
Jogo 102 - Vencedor do jogo 99 x Vencedor do jogo 100 - Atlanta, nos EUA
Sábado, 18 de julho de 2026Sábado, 18 de julho de 2026
Jogo 103 - Perdedor do jogo 101 x Perdedor do jogo 102 - Miami, nos EUA
Domingo, 19 de julho de 2026Domingo, 19 de julho de 2026
Jogo 104 - Vencedor do jogo 101 x Vencedor do jogo 102 - Nova York/Nova Jersey, nos EUA
'''

# ====== Normalização e correção de mês (jan->jun em 21/22) ======
# As páginas oficiais de partida indicam junho (ex.: 21 jun 2026; 22 jun 2026). [2](https://www.fifa.com/pt/match-centre/match/17/285023/289273/400021483)[3](https://www.fifa.com/pt/match-centre/match/17/285023/289273/400021477)[4](https://www.fifa.com/en/match-centre/match/17/285023/289273/400021494)
text = src.replace('–','-').replace('—','-')
text = text.replace('Domingo, 21 de janeiro de 2026', 'Domingo, 21 de junho de 2026')
text = text.replace('Segunda-feira, 22 de janeiro de 2026', 'Segunda-feira, 22 de junho de 2026')

months = {'janeiro':1,'fevereiro':2,'março':3,'abril':4,'maio':5,'junho':6,'julho':7,'agosto':8,'setembro':9,'outubro':10,'novembro':11,'dezembro':12}

re_date = re.compile(r'(Segunda-feira|Terça-feira|Quarta-feira|Quinta-feira|Sexta-feira|Sábado|Domingo),?\s+(\d{1,2})(?:º)?\s+de\s+([a-zç]+)\s+de\s+(\d{4})', re.IGNORECASE)
re_group = re.compile(r'Grupo\s+([A-L]):\s+(.+?)\s+x\s+(.+?)\s+(.+?)\s*-\s*(\d{1,2}h\d{2}).*?\(([^)]*)\)', re.IGNORECASE)
re_knock = re.compile(r'Jogo\s+(\d{1,3})\s*-\s*(.+?)\s+x\s+(.+?)\s*-\s*(.+)', re.IGNORECASE)
re_brt = re.compile(r'(\d{1,2}h\d{2})\s*(?:de\s*\d{1,2}\s*de\s*\w+\s*)?em\s*Brasília', re.IGNORECASE)

def parse_time(t):
    if not t: return None
    m = re.match(r'(\d{1,2})h(\d{2})', t)
    return dt.time(int(m.group(1)), int(m.group(2))) if m else None

lines = [ln.strip() for ln in text.split('\n') if ln.strip()]
current_date = None
matches = []

for ln in lines:
    dm = re_date.search(ln)
    if dm:
        d = int(dm.group(2)); mon = months[dm.group(3).lower()]; y = int(dm.group(4))
        current_date = dt.date(y, mon, d)
        continue

    gm = re_group.search(ln)
    if gm and current_date:
        grp = gm.group(1).upper()
        a = gm.group(2).strip(); b = gm.group(3).strip()
        loc = gm.group(4).strip().rstrip('-').strip()
        tloc = parse_time(gm.group(5))
        brt = re_brt.search(gm.group(6))
        tbrt = parse_time(brt.group(1)) if brt else None
        matches.append((None,'Fase de Grupos',grp,current_date,tloc,tbrt,loc,a,'',b,'',''))
        continue

    km = re_knock.search(ln)
    if km and current_date:
        mid = int(km.group(1))
        a = km.group(2).strip(); b = km.group(3).strip(); loc = km.group(4).strip()
        if 73<=mid<=88: fase='Rodada de 32 (16-avos)'
        elif 89<=mid<=96: fase='Oitavas de Final'
        elif 97<=mid<=100: fase='Quartas de Final'
        elif 101<=mid<=102: fase='Semifinais'
        elif mid==103: fase='3º Lugar'
        else: fase='Final'
        matches.append((mid,fase,'',current_date,'','',loc,a,'',b,'',''))

# IDs 1..72 para grupos
gid = 1
fixed = []
for row in matches:
    if row[0] is None:
        fixed.append((gid,)+row[1:])
        gid += 1
    else:
        fixed.append(row)

assert len(fixed)==104, f"Esperado 104 jogos; obtido {len(fixed)}"
assert gid==73, f"Esperado 72 jogos de grupos; obtido {gid-1}"

# Lista de seleções (para dropdowns)
teams = sorted({r[7] for r in fixed if r[1]=='Fase de Grupos'} | {r[9] for r in fixed if r[1]=='Fase de Grupos'})
teams = ['TBD'] + teams

# ====== Criação do Excel ======
wb = Workbook()
wb.remove(wb.active)

header_fill = PatternFill('solid', fgColor='111827')
header_font = Font(color='FFFFFF', bold=True)
center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin = Side(style='thin', color='E5E7EB')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(ws):
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
    ws.freeze_panes = 'A2'

def apply_borders(ws):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for c in row:
            c.border = border

# Regras
ws = wb.create_sheet('Regras')
ws['A1'] = 'Regras do Bolão – Copa do Mundo 2026'
ws['A1'].font = Font(bold=True, size=16)
ws['A3'] = 'Pontuação (por jogo)'
ws['A3'].font = Font(bold=True, size=12)
ws['A4'] = '• 3 pontos: placar exato'
ws['A5'] = '• 1 ponto: resultado correto (vitória/empate/derrota), sem placar exato'
ws['A6'] = '• 0 ponto: erro'
ws['A8'] = 'Prêmio'
ws['A8'].font = Font(bold=True, size=12)
ws['A9'] = '• O prêmio será um sorvete, pago pelo perdedor. Local a critério do vencedor.'
ws.column_dimensions['A'].width = 90

# Listas
wsL = wb.create_sheet('Listas')
wsL.append(['Seleções'])
wsL['A1'].font = Font(bold=True)
for t in teams: wsL.append([t])
wsL['C1'] = 'SimNao'; wsL['C1'].font = Font(bold=True)
wsL['C2'] = 'Sim'; wsL['C3'] = 'Não'
wsL.column_dimensions['A'].width = 32
wsL.column_dimensions['C'].width = 12

# Jogos
wsJ = wb.create_sheet('Jogos')
wsJ.append(['ID','Fase','Grupo','Data','Hora local','Hora (Brasília)','Local','Seleção A','Gols A','Seleção B','Gols B','Observações'])
for r in fixed:
    wsJ.append(list(r))
style_header(wsJ)
wsJ.auto_filter.ref = f"A1:L{wsJ.max_row}"
for col,w in {'A':6,'B':20,'C':8,'D':12,'E':11,'F':14,'G':30,'H':18,'I':7,'J':18,'K':7,'L':22}.items():
    wsJ.column_dimensions[col].width = w
for i in range(2, wsJ.max_row+1):
    wsJ[f'D{i}'].number_format = 'dd/mm/yyyy'
    wsJ[f'E{i}'].number_format = 'hh:mm'
    wsJ[f'F{i}'].number_format = 'hh:mm'

dv_goal = DataValidation(type='whole', operator='between', formula1='0', formula2='20', allow_blank=True)
wsJ.add_data_validation(dv_goal)
dv_goal.add(f'I2:I{wsJ.max_row}')
dv_goal.add(f'K2:K{wsJ.max_row}')

# Participantes
wsP = wb.create_sheet('Participantes')
wsP.append(['ID','Nome','Apelido','Email','WhatsApp','Pago?','Observações'])
for i in range(1,101): wsP.append([i,'','','','','Não',''])
style_header(wsP)
wsP.auto_filter.ref = "A1:G101"
for col,w in {'A':6,'B':22,'C':16,'D':28,'E':18,'F':10,'G':24}.items():
    wsP.column_dimensions[col].width = w
dv_paid = DataValidation(type='list', formula1='Listas!$C$2:$C$3', allow_blank=False)
wsP.add_data_validation(dv_paid)
dv_paid.add('F2:F101')

# Palpites
wsB = wb.create_sheet('Palpites')
wsB.append(['Participante','Jogo ID','Seleção A','Seleção B','Palpite Gols A','Palpite Gols B','Gols Reais A','Gols Reais B','Pontos'])
rows = 2500
for _ in range(rows): wsB.append(['']*9)
style_header(wsB)
wsB.auto_filter.ref = f"A1:I{wsB.max_row}"
for col,w in {'A':22,'B':9,'C':18,'D':18,'E':14,'F':14,'G':12,'H':12,'I':8}.items():
    wsB.column_dimensions[col].width = w

dv_part = DataValidation(type='list', formula1='Participantes!$B$2:$B$101', allow_blank=True)
wsB.add_data_validation(dv_part); dv_part.add(f'A2:A{wsB.max_row}')
dv_mid = DataValidation(type='list', formula1=f'Jogos!$A$2:$A${wsJ.max_row}', allow_blank=True)
wsB.add_data_validation(dv_mid); dv_mid.add(f'B2:B{wsB.max_row}')
wsB.add_data_validation(dv_goal); dv_goal.add(f'E2:E{wsB.max_row}'); dv_goal.add(f'F2:F{wsB.max_row}')

# Fórmulas (compatíveis Excel/Sheets): 3 exato, 1 resultado, 0 erro
for r in range(2, wsB.max_row+1):
    wsB[f'C{r}'] = f'=IF($B{r}="","",VLOOKUP($B{r},Jogos!$A$2:$K${wsJ.max_row},8,FALSE))'
    wsB[f'D{r}'] = f'=IF($B{r}="","",VLOOKUP($B{r},Jogos!$A$2:$K${wsJ.max_row},10,FALSE))'
    wsB[f'G{r}'] = f'=IF($B{r}="","",VLOOKUP($B{r},Jogos!$A$2:$K${wsJ.max_row},9,FALSE))'
    wsB[f'H{r}'] = f'=IF($B{r}="","",VLOOKUP($B{r},Jogos!$A$2:$K${wsJ.max_row},11,FALSE))'
    wsB[f'I{r}'] = (
        f'=IF(OR($A{r}="", $B{r}="", $E{r}="", $F{r}="", $G{r}="", $H{r}=""), "",'
        f'IF(AND($E{r}=$G{r}, $F{r}=$H{r}), 3, IF(SIGN($E{r}-$F{r})=SIGN($G{r}-$H{r}), 1, 0)))'
        f')'
    )

wsB.conditional_formatting.add(f'I2:I{wsB.max_row}', CellIsRule(operator='equal', formula=['3'], fill=PatternFill('solid', fgColor='DCFCE7')))
wsB.conditional_formatting.add(f'I2:I{wsB.max_row}', CellIsRule(operator='equal', formula=['1'], fill=PatternFill('solid', fgColor='FEF9C3')))
wsB.conditional_formatting.add(f'I2:I{wsB.max_row}', CellIsRule(operator='equal', formula=['0'], fill=PatternFill('solid', fgColor='FEE2E2')))

# Ranking
wsR = wb.create_sheet('Ranking')
wsR.append(['Participante','Pontos Totais','Acertos Exatos (3)','Resultados Corretos (>=1)','Palpites Feitos','Média'])
for i in range(2,102):
    wsR.append(['']*6)
    wsR[f'A{i}'] = f'=IF(Participantes!B{i}="","",Participantes!B{i})'
    wsR[f'B{i}'] = f'=IF($A{i}="","",SUMIF(Palpites!$A:$A,$A{i},Palpites!$I:$I))'
    wsR[f'C{i}'] = f'=IF($A{i}="","",COUNTIFS(Palpites!$A:$A,$A{i},Palpites!$I:$I,3))'
    wsR[f'D{i}'] = f'=IF($A{i}="","",COUNTIFS(Palpites!$A:$A,$A{i},Palpites!$I:$I,\">=1\"))'
    wsR[f'E{i}'] = f'=IF($A{i}="","",COUNTIF(Palpites!$A:$A,$A{i}))'
    wsR[f'F{i}'] = f'=IF(OR($A{i}="", $E{i}=0), "", $B{i}/$E{i})'
style_header(wsR)
wsR.auto_filter.ref = "A1:F101"
for col,w in {'A':24,'B':14,'C':18,'D':22,'E':14,'F':12}.items():
    wsR.column_dimensions[col].width = w
wsR['H2'] = 'Dica: ordene por Pontos Totais (desc) para ver o ranking.'
wsR['H2'].font = Font(italic=True, color='6B7280')
wsR.column_dimensions['H'].width = 65

# Acabamento: bordas e alinhamento
for wsX in [wsJ, wsP, wsB, wsR]:
    apply_borders(wsX)
    for row in wsX.iter_rows(min_row=2, max_row=wsX.max_row):
        for cell in row:
            if cell.column_letter in ['A','B','C','D','E','F','I','K']:
                cell.alignment = center
            else:
                cell.alignment = left

out = 'Bolao_Copa_2026_Calendario_Oficial.xlsx'
wb.save(out)
print('OK! Arquivo gerado:', out)
