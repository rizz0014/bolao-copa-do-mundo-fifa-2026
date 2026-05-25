import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

teams = [
    'Mexico', 'South Africa', 'South Korea', 'Czech Republic',
    'Canada', 'Bosnia and Herzegovina', 'Qatar', 'Switzerland',
    'Brazil', 'Morocco', 'Haiti', 'Scotland',
    'United States', 'Paraguay', 'Australia', 'Turkey',
    'Germany', 'Curaçao', 'Ivory Coast', 'Ecuador',
    'Netherlands', 'Japan', 'Sweden', 'Tunisia',
    'Belgium', 'Egypt', 'Iran', 'New Zealand',
    'Spain', 'Cape Verde', 'Saudi Arabia', 'Uruguay',
    'France', 'Senegal', 'Iraq', 'Norway',
    'Argentina', 'Algeria', 'Austria', 'Jordan',
    'Portugal', 'DR Congo', 'Uzbekistan', 'Colombia',
    'England', 'Croatia', 'Ghana', 'Panama',
]

match_rows = [
    (1, '2026-06-11', 'Group A', 'Mexico', 'South Africa', 'Estadio Azteca', 'UTC-6'),
    (2, '2026-06-11', 'Group A', 'South Korea', 'Czech Republic', 'Estadio Azteca', 'UTC-6'),
    (3, '2026-06-12', 'Group B', 'Canada', 'Bosnia and Herzegovina', 'BMO Field', 'UTC-4'),
    (4, '2026-06-12', 'Group D', 'United States', 'Paraguay', 'SoFi Stadium', 'UTC-7'),
    (5, '2026-06-13', 'Group C', 'Haiti', 'Scotland', 'MetLife Stadium', 'UTC-4'),
    (6, '2026-06-13', 'Group D', 'Australia', 'Turkey', 'SoFi Stadium', 'UTC-7'),
    (7, '2026-06-13', 'Group C', 'Brazil', 'Morocco', 'MetLife Stadium', 'UTC-4'),
    (8, '2026-06-13', 'Group B', 'Qatar', 'Switzerland', 'BMO Field', 'UTC-7'),
    (9, '2026-06-14', 'Group E', 'Ivory Coast', 'Ecuador', 'NRG Stadium', 'UTC-4'),
    (10, '2026-06-14', 'Group E', 'Germany', 'Curaçao', 'NRG Stadium', 'UTC-4'),
    (11, '2026-06-14', 'Group F', 'Netherlands', 'Japan', 'AT&T Stadium', 'UTC-6'),
    (12, '2026-06-14', 'Group F', 'Sweden', 'Tunisia', 'AT&T Stadium', 'UTC-6'),
    (13, '2026-06-15', 'Group H', 'Saudi Arabia', 'Uruguay', 'Mercedes-Benz Stadium', 'UTC-4'),
    (14, '2026-06-15', 'Group H', 'Spain', 'Cape Verde', 'Mercedes-Benz Stadium', 'UTC-4'),
    (15, '2026-06-15', 'Group G', 'Iran', 'New Zealand', 'Lumen Field', 'UTC-7'),
    (16, '2026-06-15', 'Group G', 'Belgium', 'Egypt', 'Lumen Field', 'UTC-7'),
    (17, '2026-06-16', 'Group I', 'France', 'Senegal', 'MetLife Stadium', 'UTC-4'),
    (18, '2026-06-16', 'Group I', 'Iraq', 'Norway', 'MetLife Stadium', 'UTC-4'),
    (19, '2026-06-16', 'Group J', 'Argentina', 'Algeria', 'AT&T Stadium', 'UTC-6'),
    (20, '2026-06-16', 'Group J', 'Austria', 'Jordan', 'AT&T Stadium', 'UTC-6'),
    (21, '2026-06-17', 'Group L', 'Ghana', 'Panama', 'AT&T Stadium', 'UTC-6'),
    (22, '2026-06-17', 'Group L', 'England', 'Croatia', 'AT&T Stadium', 'UTC-6'),
    (23, '2026-06-17', 'Group K', 'Portugal', 'DR Congo', 'Estadio Azteca', 'UTC-6'),
    (24, '2026-06-17', 'Group K', 'Uzbekistan', 'Colombia', 'Estadio Azteca', 'UTC-6'),
    (25, '2026-06-18', 'Group A', 'Czech Republic', 'South Africa', 'Estadio Akron', 'UTC-4'),
    (26, '2026-06-18', 'Group B', 'Switzerland', 'Bosnia and Herzegovina', 'Levi\'s Stadium', 'UTC-7'),
    (27, '2026-06-18', 'Group B', 'Canada', 'Qatar', 'SoFi Stadium', 'UTC-7'),
    (28, '2026-06-18', 'Group A', 'Mexico', 'South Korea', 'Mercedes-Benz Stadium', 'UTC-6'),
    (29, '2026-06-19', 'Group C', 'Brazil', 'Haiti', 'Gillette Stadium', 'UTC-4'),
    (30, '2026-06-19', 'Group C', 'Scotland', 'Morocco', 'Gillette Stadium', 'UTC-4'),
    (31, '2026-06-19', 'Group D', 'Turkey', 'Paraguay', 'BC Place', 'UTC-7'),
    (32, '2026-06-19', 'Group D', 'United States', 'Australia', 'BC Place', 'UTC-7'),
    (33, '2026-06-20', 'Group E', 'Germany', 'Ivory Coast', 'Lincoln Financial Field', 'UTC-4'),
    (34, '2026-06-20', 'Group E', 'Ecuador', 'Curaçao', 'BMO Field', 'UTC-5'),
    (35, '2026-06-20', 'Group F', 'Netherlands', 'Sweden', 'Estadio BBVA', 'UTC-5'),
    (36, '2026-06-20', 'Group F', 'Tunisia', 'Japan', 'NRG Stadium', 'UTC-6'),
    (37, '2026-06-21', 'Group D', 'Uruguay', 'Cape Verde', 'Hard Rock Stadium', 'UTC-4'),
    (38, '2026-06-21', 'Group H', 'Spain', 'Saudi Arabia', 'Mercedes-Benz Stadium', 'UTC-4'),
    (39, '2026-06-21', 'Group G', 'Belgium', 'Iran', 'SoFi Stadium', 'UTC-7'),
    (40, '2026-06-21', 'Group G', 'New Zealand', 'Egypt', 'SoFi Stadium', 'UTC-7'),
    (41, '2026-06-22', 'Group I', 'Norway', 'Senegal', 'Gillette Stadium', 'UTC-4'),
    (42, '2026-06-22', 'Group I', 'France', 'Iraq', 'Gillette Stadium', 'UTC-4'),
    (43, '2026-06-22', 'Group J', 'Argentina', 'Austria', 'Levi\'s Stadium', 'UTC-5'),
    (44, '2026-06-22', 'Group J', 'Jordan', 'Algeria', 'AT&T Stadium', 'UTC-7'),
    (45, '2026-06-23', 'Group L', 'England', 'Ghana', 'BMO Field', 'UTC-5'),
    (46, '2026-06-23', 'Group L', 'Croatia', 'Panama', 'BMO Field', 'UTC-5'),
    (47, '2026-06-23', 'Group K', 'Portugal', 'Uzbekistan', 'Estadio Azteca', 'UTC-6'),
    (48, '2026-06-23', 'Group K', 'DR Congo', 'Colombia', 'Estadio Azteca', 'UTC-6'),
    (49, '2026-06-24', 'Group C', 'Scotland', 'Brazil', 'Lincoln Financial Field', 'UTC-4'),
    (50, '2026-06-24', 'Group C', 'Morocco', 'Haiti', 'Hard Rock Stadium', 'UTC-4'),
    (51, '2026-06-24', 'Group B', 'Canada', 'Switzerland', 'BC Place', 'UTC-7'),
    (52, '2026-06-24', 'Group B', 'Bosnia and Herzegovina', 'Qatar', 'BC Place', 'UTC-7'),
    (53, '2026-06-24', 'Group A', 'Czech Republic', 'Mexico', 'Estadio Azteca', 'UTC-6'),
    (54, '2026-06-24', 'Group A', 'South Africa', 'South Korea', 'Estadio Azteca', 'UTC-6'),
    (55, '2026-06-25', 'Group E', 'Curaçao', 'Ivory Coast', 'Arrowhead Stadium', 'UTC-5'),
    (56, '2026-06-25', 'Group E', 'Germany', 'Ecuador', 'Lincoln Financial Field', 'UTC-4'),
    (57, '2026-06-25', 'Group F', 'Japan', 'Sweden', 'Levi\'s Stadium', 'UTC-5'),
    (58, '2026-06-25', 'Group F', 'Tunisia', 'Netherlands', 'Levi\'s Stadium', 'UTC-5'),
    (59, '2026-06-25', 'Group D', 'United States', 'Turkey', 'SoFi Stadium', 'UTC-7'),
    (60, '2026-06-25', 'Group D', 'Paraguay', 'Australia', 'SoFi Stadium', 'UTC-7'),
    (61, '2026-06-26', 'Group I', 'France', 'Norway', 'MetLife Stadium', 'UTC-6'),
    (62, '2026-06-26', 'Group I', 'Senegal', 'Iraq', 'MetLife Stadium', 'UTC-6'),
    (63, '2026-06-26', 'Group G', 'Egypt', 'Iran', 'BC Place', 'UTC-7'),
    (64, '2026-06-26', 'Group G', 'New Zealand', 'Belgium', 'BC Place', 'UTC-7'),
    (65, '2026-06-26', 'Group H', 'Saudi Arabia', 'Cape Verde', 'Hard Rock Stadium', 'UTC-4'),
    (66, '2026-06-26', 'Group H', 'Uruguay', 'Spain', 'NRG Stadium', 'UTC-6'),
    (67, '2026-06-27', 'Group L', 'England', 'Panama', 'BMO Field', 'UTC-4'),
    (68, '2026-06-27', 'Group L', 'Croatia', 'Ghana', 'MetLife Stadium', 'UTC-4'),
    (69, '2026-06-27', 'Group J', 'Austria', 'Algeria', 'BMO Field', 'UTC-4'),
    (70, '2026-06-27', 'Group J', 'Argentina', 'Jordan', 'AT&T Stadium', 'UTC-7'),
    (71, '2026-06-27', 'Group K', 'Colombia', 'Portugal', 'Estadio Akron', 'UTC-6'),
    (72, '2026-06-27', 'Group K', 'DR Congo', 'Uzbekistan', 'Hard Rock Stadium', 'UTC-4'),
    (73, '2026-06-28', 'Round of 32', '2A', '2B', 'SoFi Stadium', 'UTC-7'),
    (74, '2026-06-29', 'Round of 32', '1C', '2F', 'NRG Stadium', 'UTC-4'),
    (75, '2026-06-29', 'Round of 32', '1E', '3ABCDF', 'NRG Stadium', 'UTC-4'),
    (76, '2026-06-29', 'Round of 32', '1F', '2C', 'Gillette Stadium', 'UTC-6'),
    (77, '2026-06-29', 'Round of 32', '2E', '2I', 'Levi\'s Stadium', 'UTC-5'),
    (78, '2026-06-30', 'Round of 32', '1I', '3CDFGH', 'Estadio BBVA', 'UTC-5'),
    (79, '2026-06-30', 'Round of 32', '1A', '3CEFHI', 'MetLife Stadium', 'UTC-6'),
    (80, '2026-07-01', 'Round of 32', '1L', '3EHIJK', 'Estadio Azteca', 'UTC-5'),
    (81, '2026-07-01', 'Round of 32', '1D', '3BEFIJ', 'Lumen Field', 'UTC-7'),
    (82, '2026-07-01', 'Round of 32', '1G', '3AEHIJ', 'Lumen Field', 'UTC-7'),
    (83, '2026-07-02', 'Round of 32', '1H', '2J', 'SoFi Stadium', 'UTC-4'),
    (84, '2026-07-02', 'Round of 32', '2K', '2L', 'Levi\'s Stadium', 'UTC-7'),
    (85, '2026-07-03', 'Round of 32', '1B', '3EFGIJ', 'AT&T Stadium', 'UTC-4'),
    (86, '2026-07-03', 'Round of 32', '2D', '2G', 'Hard Rock Stadium', 'UTC-5'),
    (87, '2026-07-03', 'Round of 32', '1J', '2H', 'AT&T Stadium', 'UTC-4'),
    (88, '2026-07-03', 'Round of 32', '1K', '3DEIJL', 'AT&T Stadium', 'UTC-4'),
    (89, '2026-07-04', 'Round of 16', 'W74', 'W77', 'NRG Stadium', 'UTC-4'),
    (90, '2026-07-04', 'Round of 16', 'W73', 'W75', 'NRG Stadium', 'UTC-4'),
    (91, '2026-07-05', 'Round of 16', 'W76', 'W78', 'Lincoln Financial Field', 'UTC-4'),
    (92, '2026-07-05', 'Round of 16', 'W79', 'W80', 'MetLife Stadium', 'UTC-6'),
    (93, '2026-07-06', 'Round of 16', 'W83', 'W84', 'Estadio Azteca', 'UTC-5'),
    (94, '2026-07-06', 'Round of 16', 'W81', 'W82', 'AT&T Stadium', 'UTC-6'),
    (95, '2026-07-07', 'Round of 16', 'W86', 'W88', 'Lumen Field', 'UTC-7'),
    (96, '2026-07-07', 'Round of 16', 'W85', 'W87', 'BC Place', 'UTC-7'),
    (97, '2026-07-09', 'Quarterfinal', 'W89', 'W90', 'Gillette Stadium', 'UTC-7'),
    (98, '2026-07-10', 'Quarterfinal', 'W93', 'W94', 'SoFi Stadium', 'UTC-4'),
    (99, '2026-07-11', 'Quarterfinal', 'W91', 'W92', 'Hard Rock Stadium', 'UTC-5'),
    (100, '2026-07-11', 'Quarterfinal', 'W95', 'W96', 'Arrowhead Stadium', 'UTC-5'),
    (101, '2026-07-14', 'Semifinal', 'W97', 'W98', 'AT&T Stadium', 'UTC-5'),
    (102, '2026-07-15', 'Semifinal', 'W99', 'W100', 'AT&T Stadium', 'UTC-5'),
    (103, '2026-07-18', 'Third-place', 'RU101', 'RU102', 'Hard Rock Stadium', 'UTC-5'),
    (104, '2026-07-19', 'Final', 'W101', 'W102', 'MetLife Stadium', 'UTC-4'),
]

wb = Workbook()
ws_games = wb.active
ws_games.title = 'Tabela de Jogos'
ws_lists = wb.create_sheet('Lists')
ws_participantes = wb.create_sheet('Cadastro de Participantes')
ws_palpites = wb.create_sheet('Controle de Palpites')
ws_rank = wb.create_sheet('Classificação e Ranking Geral')
ws_painel = wb.create_sheet('Painel Geral')

header_fill = PatternFill(start_color='FFDDEBF7', end_color='FFDDEBF7', fill_type='solid')
header_font = Font(bold=True)
center = Alignment(horizontal='center', vertical='center')

# Fill Lists sheet for dropdown references
ws_lists.append(['Team Names', 'Match Number'])
for team in teams:
    ws_lists.append([team, None])
for i in range(1, 105):
    ws_lists.cell(row=i+1, column=2, value=i)

# Tabela de Jogos
columns = ['Match #', 'Date', 'Stage', 'Home Team', 'Away Team', 'Venue', 'Time Zone', 'Actual Home', 'Actual Away', 'Result Notes']
ws_games.append(columns)
for cell in ws_games[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
for row in match_rows:
    ws_games.append(list(row) + ['', '', ''])
for col, width in enumerate([10, 14, 18, 22, 22, 28, 10, 12, 12, 30], 1):
    ws_games.column_dimensions[get_column_letter(col)].width = width

# Cadastro de Participantes
participant_headers = ['Participant', 'Email', 'Phone', 'Nickname', 'Group / Team']
ws_participantes.append(participant_headers)
for cell in ws_participantes[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
for col, width in enumerate([24, 32, 18, 18, 22], 1):
    ws_participantes.column_dimensions[get_column_letter(col)].width = width

# Controle de Palpites
palpites_headers = ['Row', 'Participant', 'Match #', 'Home Team', 'Away Team', 'Pred Home', 'Pred Away', 'Real Home', 'Real Away', 'Points', 'Outcome Status']
ws_palpites.append(palpites_headers)
for cell in ws_palpites[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
for col, width in enumerate([6, 24, 10, 20, 20, 12, 12, 12, 12, 12, 18], 1):
    ws_palpites.column_dimensions[get_column_letter(col)].width = width
for i in range(2, 202):
    ws_palpites.cell(row=i, column=1, value=i-1)
    ws_palpites.cell(row=i, column=4).alignment = center
    ws_palpites.cell(row=i, column=5).alignment = center
    ws_palpites.cell(row=i, column=6).alignment = center
    ws_palpites.cell(row=i, column=7).alignment = center
    ws_palpites.cell(row=i, column=10).alignment = center
    ws_palpites.cell(row=i, column=11).alignment = center
    ws_palpites.cell(row=i, column=4, value=f'=IFERROR(VLOOKUP($C{i},\'Tabela de Jogos\'!$A:$J,4,FALSE),"")')
    ws_palpites.cell(row=i, column=5, value=f'=IFERROR(VLOOKUP($C{i},\'Tabela de Jogos\'!$A:$J,5,FALSE),"")')
    ws_palpites.cell(row=i, column=8, value=f'=IFERROR(VLOOKUP($C{i},\'Tabela de Jogos\'!$A:$J,8,FALSE),"")')
    ws_palpites.cell(row=i, column=9, value=f'=IFERROR(VLOOKUP($C{i},\'Tabela de Jogos\'!$A:$J,9,FALSE),"")')
    ws_palpites.cell(row=i, column=10, value=f'=IF(OR($F{i}="",$G{i}="",$H{i}="",$I{i}=""),"",IF(AND($F{i}=$H{i},$G{i}=$I{i}),3,IF(SIGN($F{i}-$G{i})=SIGN($H{i}-$I{i}),1,0)))')
    ws_palpites.cell(row=i, column=11, value=f'=IF($J{i}="","",IF($J{i}=3,"Exact","IF($J{i}=1,"Correct","Miss")))')

# Data validation for match number and participant
match_validation = DataValidation(type='list', formula1='=Lists!$B$2:$B$105', allow_blank=True)
participant_validation = DataValidation(type='list', formula1='=Cadastro de Participantes!$A$2:$A$100', allow_blank=True)
ws_palpites.add_data_validation(match_validation)
ws_palpites.add_data_validation(participant_validation)
match_validation.add(f'C2:C201')
participant_validation.add(f'B2:B201')

# Formatting for Tabela de Jogos
for row in ws_games.iter_rows(min_row=2, max_row=105, min_col=1, max_col=10):
    for cell in row:
        cell.alignment = center

# Classificação e Ranking Geral
rank_headers = ['Participant', 'Matches', 'Exact Hits', 'Correct Outcomes', 'Total Points', 'Avg Points', 'Rank']
ws_rank.append(rank_headers)
for cell in ws_rank[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
for col, width in enumerate([24, 12, 14, 16, 14, 14, 8], 1):
    ws_rank.column_dimensions[get_column_letter(col)].width = width
for i in range(2, 102):
    ws_rank.cell(row=i, column=1, value=f'=IFERROR(INDIRECT("Cadastro de Participantes!A" & ROW()),"")')
    ws_rank.cell(row=i, column=2, value=f'=IF($A{i}="","",COUNTIF(\'Controle de Palpites\'!$B:$B,$A{i}))')
    ws_rank.cell(row=i, column=3, value=f'=IF($A{i}="","",SUMIFS(\'Controle de Palpites\'!$J:$J,\'Controle de Palpites\'!$B:$B,$A{i},\'Controle de Palpites\'!$J:$J,3))')
    ws_rank.cell(row=i, column=4, value=f'=IF($A{i}="","",SUMIFS(\'Controle de Palpites\'!$K:$K,\'Controle de Palpites\'!$B:$B,$A{i}))')
    ws_rank.cell(row=i, column=5, value=f'=IF($A{i}="","",SUMIFS(\'Controle de Palpites\'!$J:$J,\'Controle de Palpites\'!$B:$B,$A{i}))')
    ws_rank.cell(row=i, column=6, value=f'=IF($B{i}=0,"",ROUND($E{i}/$B{i},2))')
    ws_rank.cell(row=i, column=7, value=f'=IF($E{i}="","",RANK($E{i},$E$2:$E$101,0)+COUNTIFS($E$2:$E$101,$E{i},$C$2:$C$101,">"&$C{i}))')

# Painel Geral
ws_painel.append(['Painel Geral do Bolão'])
ws_painel.append(['Última atualização: preencha resultados em Tabela de Jogos e palpites em Controle de Palpites.'])
ws_painel.append(['Prêmio: sorvete pago pelo perdedor, local a critério do vencedor.'])
ws_painel.append([])
ws_painel.append(['Top 5 Participantes'])
for i in range(1, 6):
    ws_painel.append([f'{i}º', f'=IFERROR(INDEX(\'Classificação e Ranking Geral\'!$A:$A, MATCH({i}, \'Classificação e Ranking Geral\'!$G:$G, 0)), "")',
                     f'=IFERROR(INDEX(\'Classificação e Ranking Geral\'!$E:$E, MATCH({i}, \'Classificação e Ranking Geral\'!$G:$G, 0)), "")'])
for col, width in enumerate([8, 24, 14], 1):
    ws_painel.column_dimensions[get_column_letter(col)].width = width
for row in ws_painel.iter_rows(min_row=1, max_row=20, min_col=1, max_col=3):
    for cell in row:
        if cell.value is not None:
            cell.alignment = Alignment(horizontal='left', vertical='center')

# Hide helper sheet
ws_lists.sheet_state = 'hidden'

wb.save('bolao_copa_2026.xlsx')
